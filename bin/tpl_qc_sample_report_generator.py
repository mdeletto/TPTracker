#!/usr/bin/python2.7

from BeautifulSoup import BeautifulSoup as parse_html_string
import openpyxl
import string
import re
import os
import shutil
import cgi
import sys
from collections import defaultdict

print "Content-type: text/html\n\n"

def write_html_output():

    HTML_string = ""

    def print_html_header(HTML_string):

        HTML_string += "<html>"

        return HTML_string

    def print_CSS(HTML_string):
        """Takes the HTML string passed to it and adds on the CSS style section."""
        
        HTML_string += """
         
        <style type="text/css">
        
            h1 {
                color:#2D4782;
            }
        
            hr {
                border-color: #2D4782;
                width: 80%;
            }
            
            .page
            {   
                size: landscape;
                margin-left:auto; 
                margin-right:auto;
                display: block;
                /*
                border: solid;
                border-color:#2D4782;
                */
                padding-left: 5px;
                padding-top: 5px;
                padding-bottom: 5px;
                padding-right: 5px;
            }
            
            .Table
            {
                font-family: Arial, Helvetica, sans-serif;
                color:#2D4782;
                display: table;
                font-size: 80%;
                margin-left:auto; 
                margin-right:auto;
                padding-left: 5px;
                padding-top: 5px;
                padding-bottom: 5px;
                padding-right: 5px;

            }

            .Inline-Table
            {
                font-family: Arial, Helvetica, sans-serif;
                color:#2D4782;
                display: table;   
                display: inline-block;
                font-size: 150%;    
                margin-left:auto; 
                margin-right:auto;
                padding-left: 5px;
                padding-top: 5px;
                padding-right: 5px;
                vertical-align:top;
                
        
            }

            .Title
            {
                display: table-caption;
                text-align: center;
                font-weight: bold;
                font-size: 150%;
                
            }
            .Heading
            {
                display: table-row;
                font-weight: bold;
                text-align: center;
                background-color: #2D4782;
                color:white;
            }
            .Row
            {
                display: table-row;
            }
            
            
            .Row-no-left
            {
                display: table-row;
                float:left;
                word-wrap: break-word;
                width:800px;
            }
            
            .Cell
            {
                display: table-cell;
                border: solid #2D4782;
                vertical-align:middle;
                margin-left:auto; 
                margin-right:auto;
                border-width: thin;
                padding-left: 5px;
                padding-right: 5px;
            }

            .Cell-bold-and-red
            {
                display: table-cell;
                border: solid #2D4782;
                font-weight: bold;
                color: red;
                vertical-align:middle;
                margin-left:auto; 
                margin-right:auto;
                border-width: thin;
                padding-left: 5px;
                padding-right: 5px;
            }

            .Cell-as-header
            {
                display: table-cell;
                font-weight: bold;
                background-color: #2D4782;
                color:white;
                border: solid #2D4782;
                vertical-align:middle;
                margin-left:auto; 
                margin-right:auto;
                border-width: thin;
                padding-left: 5px;
                padding-right: 5px;
            }
            
            .Cell-as-header-no-bold
            {
                display: table-cell;
                background-color: #2D4782;
                color:white;
                border: solid #2D4782;
                vertical-align:middle;
                margin-left:auto; 
                margin-right:auto;
                border-width: thin;
                padding-left: 5px;
                padding-right: 5px;
            }

            .Cell-no-border
            {
                display: table-cell;
                border: none;
                vertical-align:middle;
                margin-left:auto; 
                margin-right:auto;
                padding-left: 5px;
                padding-right: 5px;
            }
            
            .Cell-no-border-whole-row
            {
                text-align:left;
                
                display: table-cell;
                border: none;
                float:left;
                width:100%;
                vertical-align:middle;
                padding-left: 5px;
                padding-right: 10px;
            }

            .Cell-dotted-border
            {
                display: table-cell;
                border: dotted #2D4782;
                vertical-align:middle;
                margin-left:auto; 
                margin-right:auto;
                padding-left: 5px;
                padding-right: 5px;
            }
            
            .img-header

            {
                display: block;
                align:middle;
                margin-left:auto;
                margin-right:auto;
                max-width:100%;
                max-height:100%;
                padding:1px;
                border:none;
            
            }   
            
            
            
        </style>
        </head>
        <title>TPL QC Sample Report Generator</title>
        <body>
        <div class="page">
        """
        
        return HTML_string
    
    
    
    def print_cell_in_div_table(line, css_class):
        tmp_str = ''
        tmp_str += '<div class="%s">' % css_class
        tmp_str += '%s' % line
        tmp_str += '</div>'
        
        return tmp_str
    
    def print_bold_cell_in_div_table(line):
        tmp_str = ''
        tmp_str += '<div class="Cell" style="font-weight: bold;">'
        tmp_str += '%s' % line
        tmp_str += '</div>'
        
        return tmp_str
        


    def print_qc_variants(HTML_string):

        # open workbook and grab 'qc' sheet
        
        wb = openpyxl.load_workbook('/var/www/TPL/ftp/QC/spreadsheets/%s' % input_spreadsheet)
        qc_sheet = wb.get_sheet_by_name("qc")
        
        # initialize dict counters.  This will be used to group variants into categories
        
        comment_category_counter = defaultdict(int)
        pass_or_fail_category_counter = defaultdict(int)
        
        # determine max row to set rows to iterate over
        
        max_row = qc_sheet.max_row

        # begin looping over rows
        values_to_print = []
        
        row_counter = 1
        for row in qc_sheet.iter_rows('A1:Q%s' % (str(max_row))):
            if row_counter == 1:
                pass
            else:
                
                class ColumnCategories:
                    def __init__(self, row):
                        self.comment = qc_sheet.cell(row = row_counter, column = 1)
                        self.pass_or_fail = qc_sheet.cell(row = row_counter, column = 2)
                        self.chr = qc_sheet.cell(row = row_counter, column = 3)
                        self.gene = qc_sheet.cell(row = row_counter, column = 4)
                        self.pos = qc_sheet.cell(row = row_counter, column = 5)
                        self.ref = qc_sheet.cell(row = row_counter, column = 6)
                        self.alt = qc_sheet.cell(row = row_counter, column = 7)
                        self.mutation_type = qc_sheet.cell(row = row_counter, column = 8)
                        self.status = qc_sheet.cell(row = row_counter, column = 9)
                        self.main_consequence = qc_sheet.cell(row = row_counter, column = 10)
                        self.t_vaf = qc_sheet.cell(row = row_counter, column = 11)
                        self.n_vaf = qc_sheet.cell(row = row_counter, column = 12)
                        self.expected_vaf = qc_sheet.cell(row = row_counter, column = 13)
                        self.expected_hgvsc = qc_sheet.cell(row = row_counter, column = 15)
                        self.expected_hgvsp = qc_sheet.cell(row = row_counter, column = 16)
                        self.mutation_ID = qc_sheet.cell(row = row_counter, column = 17)
        
                
                column_in_qc_sheet = ColumnCategories(row)
                
                # parse out position.  Its trapped awkwardly in a hyperlink
                
                if re.search("HYPERLINK", column_in_qc_sheet.pos.value):
                    match = re.search('"(\d+?)"', column_in_qc_sheet.pos.value)
                    pos = match.group(1)
                else:
                    pos = column_in_qc_sheet.pos.value
                
                try:
                    t_vaf = "{0:.1f}%".format(float(column_in_qc_sheet.t_vaf.value) * 100),
                except:
                    t_vaf = ""
                    
                tmp_values_to_print = [column_in_qc_sheet.comment.value,
                                       column_in_qc_sheet.pass_or_fail.value,
                                       column_in_qc_sheet.chr.value + ":" + str(pos),
                                       column_in_qc_sheet.gene.value,
                                       #column_in_qc_sheet.ref.value + "/" + column_in_qc_sheet.alt.value,
                                       column_in_qc_sheet.mutation_type.value,
                                       column_in_qc_sheet.status.value,
                                       column_in_qc_sheet.main_consequence.value,
                                       t_vaf,
                                       column_in_qc_sheet.expected_vaf.value,
                                       column_in_qc_sheet.expected_hgvsc.value,
                                       column_in_qc_sheet.expected_hgvsp.value
                                       ]
                tmp_values_to_print = map(str, tmp_values_to_print)
                values_to_print.append(tmp_values_to_print)
                
                split_comment = column_in_qc_sheet.comment.value.split(":")
                #print "\t".join(tmp_values_to_print)
                comment_category = split_comment[0]
                comment_category_counter[comment_category] += 1
                pass_or_fail_category_counter[column_in_qc_sheet.pass_or_fail.value.strip()] += 1
            
            row_counter += 1

        ### CALCULATE TOTAL IN PASS_OR_FAIL_DICT
        
        for k in pass_or_fail_category_counter.keys():
            # don't include variants are filtered by the pipeline in calculations
            if k == "FILTERED":
                pass
            else:
                pass_or_fail_category_counter['TOTAL VARIANTS'] += pass_or_fail_category_counter[k]

        ### PRINT VARIANT CATEGORY SUMMARY ###
        
        HTML_string += '<div>'
        HTML_string += '<p style="width:80%;margin:auto;color:#2D4782;">'
        HTML_string += """<b><u>METHODS:</u></b><br><br>Variants were detected using the latest version of the Targeted NGS Pipeline (please consult SOP for details).  All variant coordinates refer to the GRCh37/hg19 assembly.  Expected allele frequencies are based on resources from Thermofisher Scientific for the Acrometrix Oncology Hotspot Control.<br><br>
                        Variants are classified into the following detection categories:<br>
                        <ol style="width:80%;margin:auto;color:#2D4782;">
                            <li>PASS - Detected and within expected VAF interval</li>
                            <li>PASS-W - Detected, but not within expected VAF interval</li>
                            <li>FAIL - No evidence of variant in sequencing reads</li>
                        </ol>
                        """
        HTML_string += '</p>'
        HTML_string += '<hr>'
        HTML_string += '<p style="width:80%;margin:auto;color:#2D4782;">'
        HTML_string += """<b><u>TARGET REGIONS:</u></b> %s<br><br>""" % panel
        HTML_string += '</p>'
        HTML_string += '<hr>'
        HTML_string += '<p style="width:80%;margin:auto;color:#2D4782;">'
        HTML_string += """<b><u>REPORT NAME:</u></b> %s<br><br>""" % basename
        HTML_string += '</p>'
        HTML_string += '<hr>'
        HTML_string += '<p style="width:80%;margin:auto;color:#2D4782;">'
        HTML_string += """<b><u>COMMENTS:</u></b><br><br>"""
        HTML_string += """%s""" % text_comment
        HTML_string += '</p>'
        HTML_string += '</div>'
        
        HTML_string += '<hr>'


        # Keep the next two tables inline with each other
        HTML_string += '<div class="Table">'

        # Generate VARIANT SUMMARY table
        HTML_string += '<div class="Inline-Table">'
        HTML_string += '<div class="Table">'
        HTML_string += '<div style="font-size: 120%;" class="Title">'
        HTML_string += 'VARIANT COUNTS'
        HTML_string += '</div>'
        
        for header in sorted(pass_or_fail_category_counter.keys()):
            HTML_string += '<div class="Row">'
            tmp_list = [header, pass_or_fail_category_counter[header]]
            for value in tmp_list:
                if value == header:
                    if re.search("TOTAL", value):
                        HTML_string += print_cell_in_div_table(value, "Cell-as-header")
                    else:
                        HTML_string += print_cell_in_div_table(value, "Cell-as-header-no-bold")
                else:
                    HTML_string += print_cell_in_div_table(value, "Cell")
            HTML_string += '</div>'

        HTML_string += '</div>'
        HTML_string += '</div>'


        # Generate VARIANT SUMMARY table
        HTML_string += '<div class="Inline-Table">'
        HTML_string += '<div class="Table">'
        HTML_string += '<div style="font-size: 120%;" class="Title">'
        HTML_string += 'PERCENT DETECTED'
        HTML_string += '</div>'
        
        category_to_percent = {
                               'PASS + PASS-W' : "{0:.2f}%".format( ((float(pass_or_fail_category_counter['PASS']) + float(pass_or_fail_category_counter['PASS-W'])) / float(pass_or_fail_category_counter['TOTAL VARIANTS']))  * 100),
                               'FAIL' : "{0:.2f}%".format( (float(pass_or_fail_category_counter['FAIL']) / float(pass_or_fail_category_counter['TOTAL VARIANTS']))  * 100)
                               }
        
        for header in sorted(category_to_percent.keys()):
            HTML_string += '<div class="Row">'
            tmp_list = [header, category_to_percent[header]]
            for value in tmp_list:
                if value == header:
                    HTML_string += print_cell_in_div_table(value, "Cell-as-header")
                else:
                    HTML_string += print_cell_in_div_table(value, "Cell")
            HTML_string += '</div>'
        
        HTML_string += '</div>'
        HTML_string += '</div>'

        
        # Close center div
        
        HTML_string += '</div>'

        # Separate section
        
        HTML_string += '<hr>'

        ### PRINT VARIANT TABLE ###

        HTML_string += '<div style="width:80%;" class="Table">'
        HTML_string += '<div class="Title">'
        HTML_string += 'VARIANTS ASSESSED'
        HTML_string += '</div>'
        
        # define headers to print
        
        headers_to_print = ["COMMENT",
                            "PASS/FAIL",
                            "CHR:POS",
                            "GENE",
                            #"REF/ALT",
                            "MUTATION TYPE",
                            "STATUS",
                            "MAIN CONSEQUENCE",
                            "VAF",
                            "EXPECTED VAF",
                            "EXPECTED HGVSC",
                            "EXPECTED HGVSP"]
        
        HTML_string += '<div class="Heading">'
        for header in headers_to_print:
            HTML_string += print_cell_in_div_table(header, "Cell")
        HTML_string += '</div>'   
        
        # print variants
        for entry in values_to_print:
            HTML_string += '<div class="Row">'
            for value in entry:
                # We don't want to print None values to the table
                if value is None or re.search("None",value):
                    value = ""
                if re.search("FAIL", value):
                    HTML_string += print_cell_in_div_table(value, "Cell-bold-and-red")
                else:
                    HTML_string += print_cell_in_div_table(value, "Cell")
            HTML_string += '</div>'
            
        HTML_string += '</div>'
        
        HTML_string += '</div>'
        
        return HTML_string



    def print_additional_comments(HTML_string):
        
        HTML_string += '<div class="Table" style="text-align:left;border:none;">'
        
        HTML_string += '<div class="Row-no-left">'
        #HTML_string += print_cell_in_div_table("Additional Comments:", "Cell-as-header-no-bold")
        HTML_string += print_cell_in_div_table('<b>ADDITIONAL COMMENTS: </b><br><br>'+text_comment, "Cell-no-border-whole-row")
        HTML_string += '</div>'
        
        HTML_string += '</div>'
        
        return HTML_string

    def print_html_footer(HTML_string):
        
        HTML_string += '</div'
        HTML_string += '</body>'
        HTML_string += '</html>'

        return HTML_string

    
    HTML_string = print_html_header(HTML_string)
    HTML_string = print_CSS(HTML_string)
    HTML_string += '<img class="img-header" src="/TPL/images/tpl_report_header.png" alt="logo" />'
    HTML_string += '<h1 style="text-align:center;">QC SAMPLE VARIANT REPORT</h1>'
    HTML_string += '<hr>'
    HTML_string = print_qc_variants(HTML_string)
    HTML_string = print_html_footer(HTML_string)
    pretty_HTML_string = parse_html_string(HTML_string).prettify()

    return pretty_HTML_string

def column_to_number(c):
    """Return number corresponding to excel-style column."""
    number=-25
    for l in c:
        if not l in string.ascii_letters:
            return False
        number+=ord(l.upper())-64+25
    return number


def set_output_dir(runName):
    """Takes runName as input and creates a new output directory.  The directory is overwritten if it exists already."""
    os.chdir("/var/www/TPL/cgi-bin/output/QCReportGenerator")
    output_dir = "%s/%s" % (os.getcwd(), runName)
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir)
    os.chdir("%s/%s" % (os.getcwd(), runName))

def gather_HTML_form_values(form):
            
    basename = form.getvalue('basename')
    basename = re.sub(" ", "_", basename.strip())
    panel = form.getvalue('panel')
    input_spreadsheet = form.getvalue('input_spreadsheet')

    if form.getvalue('comment'):
        text_comment = form.getvalue('comment')
    else:
        text_comment = False

    return_values = [basename,
                     panel,
                     input_spreadsheet,
                     text_comment]

    for value in return_values:
        if value is None:
            print "Mandatory value not defined"
            sys.exit(1)

    return return_values



# Get CGI form values
form = cgi.FieldStorage()
basename, panel, input_spreadsheet, text_comment = (i for i in gather_HTML_form_values(form))

# Set output dir
set_output_dir(basename)

print write_html_output()
html_output = open("%s/%s.html" % (os.getcwd(), basename), "w")
html_output.write(write_html_output())